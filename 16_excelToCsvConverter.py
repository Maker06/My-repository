
import openpyxl, csv, os, re

import logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s -  %(levelname)s -  %(message)s')
logging.disable(logging.CRITICAL)

#for excelFile in os.listdir('.'):l
for excelFile in os.listdir('c:\\Users\\Mykhailo Maksymiuk\\PycharmProjects\\AutomateBoringStuff'):
    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        csvFileName = re.compile(r'(.*)(\.xlsx)').search(excelFile).group(1) + '_' + sheet.title + '.csv'
        # logging.debug('csvFileName: ' + str(csvFileName))
        csvFile = open(csvFileName, 'w', newline='')
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFile)
        # logging.debug('csvOutput: ' + str(csvWriter))
        # Loop through every row in the sheet
        for rowNum in sheet.rows:
            rowData = []
            logging.debug('rowNum: ' + str(rowNum))
            # append each cell to this list
            # Loop through each cell in the row.
            for colNum in rowNum:
                # Append each cell's data to rowData.
                logging.debug('colNum: ' + str(colNum))
                # cellData = sheet.cell(row=rowNum, column=colNum).value
                rowData.append(colNum.value)
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFile.close()