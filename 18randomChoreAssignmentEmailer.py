#! python3
# 18randomChoreAssignmentEmailer.py - Sends emails with random chores

import openpyxl, smtplib, sys, random, pathlib
from pathlib import Path
# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('randomChoreEmailer.xlsx')
sheet = wb['Sheet1']

chores = ['dishes', 'bathroom', 'vacuum', 'walk dog', 'dow', 'yoohoo!']
x = round(len(chores)/sheet.max_row)

# counting number of times program was running using log file
log = open('c:\\Users\\Mykhailo Maksymiuk\\PycharmProjects\\AutomateBoringStuff\\log.txt')
count = int(log.read())
print(count)
# Log in to email account.
'''smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
#smtpObj.login('testpythonskills@gmail.com', sys.argv[1])
smtpObj.login('testpythonskills@gmail.com', 'vYWtS8KgR5kyEcE')'''

# Send out emails with chores.
for r in range(2, sheet.max_row + 1):
    name = sheet.cell(row=r, column=1).value
    email = sheet.cell(row=r, column=2).value
    randomChore = random.choice(chores)
    chores.remove(randomChore)  # this chore is now taken, so remove it
    for col in range (3+count, 3+x+count):
        if randomChore == sheet.cell(row=r, column=col).value:
            continue
        else:
            sheet.cell(row=r, column=col).value = randomChore
    wb.save('randomChoreEmailer.xlsx')

    '''body = "Subject: %s for this week.\nDear %s,\nYour chore is %s." % (randomChore, name, randomChore)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('testpythonskills@gmail.com', email, body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()'''
count+=1
log = open('c:\\Users\\Mykhailo Maksymiuk\\PycharmProjects\\AutomateBoringStuff\\log.txt', 'w')
log.write(str(count))
log.close()
print(count)
