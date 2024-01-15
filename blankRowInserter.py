#! python3
# multiplicationTable.py - Starting at row N, the program inserts M blank rows into the spreadsheet.

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('duesRecords.xlsx')
old_sheet = wb.get_sheet_by_name('Sheet1')
old_sheet.title = 'Sheet1.5'

wb.create_sheet('Sheet1', 0)
new_sheet = wb.get_sheet_by_name('Sheet1')

N = int(input('Enter the number of row: '))
M = int(input('Enter the number of empty rows: '))

max_row = old_sheet.max_row
max_col = old_sheet.max_column

for col_num in range(1, max_col):
    for row_num in range(1, N):
        new_sheet.cell(row=row_num, column=col_num).value = old_sheet.cell(row=row_num, column=col_num).value

for row_num in range(N, N+M):
    for col_num in range(1, max_col):
        new_sheet.cell(row=row_num, column=col_num).value = ''
        new_sheet.cell(row=row_num+M, column=col_num).value = old_sheet.cell(row=row_num, column=col_num).value

for row_num in range(N+M, max_row+M+1):
    for col_num in range(1, max_col):
        new_sheet.cell(row=row_num, column=col_num).value = old_sheet.cell(row=row_num-M, column=col_num).value

wb.save('copy_duesRecords.xlsx')